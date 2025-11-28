from flask import Flask, render_template, request, redirect, url_for, session
from models import Room, Customer, Booking, Hotel, Employee
from datetime import datetime, date
import openpyxl, os

app = Flask(__name__)
app.secret_key = "securekey123"

# -------------------------------------
# Create Excel file if not exists
# -------------------------------------
if not os.path.exists("data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["BookingID","Name","Email","Phone","RoomType","Guests","Total","Status"])
    wb.save("data.xlsx")

# -------------------------------------
# Hotel and Room Setup
# -------------------------------------
hotel = Hotel("Nile View Hotel","Cairo, Egypt","+20 100 555 7777")
hotel.addRoom(Room(1,"Single Room",500,"Cozy single room with Wi-Fi",maxPeople=1,roomQuantity=5))
hotel.addRoom(Room(2,"Double Room",800,"Spacious room ideal for couples",maxPeople=2,roomQuantity=3))
hotel.addRoom(Room(3,"Deluxe Suite",1500,"Luxury suite with Nile view",maxPeople=4,roomQuantity=2))

admin = Employee(1,"Admin","Manager","admin","1234")

# -------------------------------------
# Routes
# -------------------------------------
@app.route('/')
def home():
    return render_template('home.html', rooms=hotel.displayAvailableRooms(), hotel=hotel)

@app.route('/about')
def about():
    return render_template('about.html', hotel=hotel)

# -------------------------------------
# Booking Page
# -------------------------------------
@app.route('/book/<int:room_id>', methods=['GET','POST'])
def book(room_id):
    room = next((r for r in hotel.listOfRooms if r.roomNumber==room_id), None)
    if not room:
        return "Room not found",404

    error=None
    if request.method=='POST':
        name=request.form['name'].strip()
        email=request.form['email'].strip()
        phone=request.form['phone'].strip()
        guests_str=request.form['guests']
        check_in_str=request.form['check_in']
        check_out_str=request.form['check_out']

        if not name or not email or not phone or not guests_str:
            error="⚠️ Please fill in all fields."
        else:
            try:
                guests=int(guests_str)
                check_in=datetime.strptime(check_in_str,'%Y-%m-%d').date()
                check_out=datetime.strptime(check_out_str,'%Y-%m-%d').date()

                if guests>room.maxPeople:
                    error=f"⚠️ {room.roomType} can host only {room.maxPeople} guest(s)."
                elif check_in<date.today():
                    error="⚠️ Check-in date cannot be in the past."
                elif check_out<=check_in:
                    error="⚠️ Check-out date must be after check-in date."
                elif not room.checkAvailability():
                    error=f"⚠️ Sorry, all {room.roomType}s are fully booked."
                else:
                    booking_id=datetime.now().strftime("%Y%m%d%H%M%S")
                    customer=Customer(booking_id,name,email,phone)
                    booking=Booking(booking_id,customer,room,datetime.combine(check_in,datetime.min.time()),datetime.combine(check_out,datetime.min.time()))
                    total=booking.calculateTotalAmount()
                    room.bookRoom()

                    wb=openpyxl.load_workbook('data.xlsx')
                    ws=wb.active
                    ws.append([booking.bookingID,name,email,phone,room.roomType,guests,total,booking.bookingStatus])
                    wb.save('data.xlsx')

                    return render_template('success.html',name=name,total=total)
            except ValueError:
                error="⚠️ Invalid input. Please check your dates and number of guests."

    return render_template('booking.html',room=room,hotel=hotel,error=error)


# -------------------------------------
# Admin Login
# -------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            error = "⚠️ Please enter both username and password."
        elif admin.login(username, password):
            session['admin'] = True
            return redirect(url_for('admin_panel'))
        else:
            error = "❌ Wrong credentials! Try again."

    return render_template('login.html', error=error)

# -------------------------------------
# Admin Logout
# -------------------------------------
@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('home'))

# -------------------------------------
# Admin Panel (Bookings Table)
# -------------------------------------
@app.route('/admin')
def admin_panel():
    if 'admin' not in session:
        return redirect(url_for('login'))

    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    bookings = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]

    return render_template('admin.html', bookings=bookings, hotel=hotel)


if __name__ == "__main__":
    app.run(debug=True)